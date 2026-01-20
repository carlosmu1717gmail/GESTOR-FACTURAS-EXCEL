import os
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Variable global para mantener el workbook y contador
_workbook = None
_sheet = None
_current_row_number = 1

def inicializar_excel(filename: str = "facturas.xlsx"):
    """
    Inicializa un nuevo archivo Excel con encabezados.
    """
    global _workbook, _sheet, _current_row_number
    
    columns_spanish = [
        "#", "Nombre Archivo", "Fecha Expedición",
        "Número Factura", "Nombre", "NIF",
        "Base Imponible", "% IVA", "Cuota IVA", "Total Factura",
        "%Retención", "Retenciones", "Forma de Pago",
        "Observaciones", "Confianza"
    ]
    
    _workbook = Workbook()
    _sheet = _workbook.active
    _sheet.title = "FACTURAS"
    
    # Escribir encabezados
    _sheet.append(columns_spanish)
    
    # Aplicar formato a encabezados: fondo azul oscuro, texto blanco, negrita
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in _sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Insertar 2 líneas en blanco
    _sheet.append([""] * len(columns_spanish))
    _sheet.append([""] * len(columns_spanish))
    
    _current_row_number = 1

def save_to_excel(data: dict, filename: str = "facturas.xlsx"):
    """
    Guarda los datos de la factura en el Excel en memoria.
    """
    global _workbook, _sheet, _current_row_number
    
    if _workbook is None:
        inicializar_excel(filename)
    
    columns = [
        "file_name", "fecha_expedicion", "numero_factura",
        "contraparte_nombre", "contraparte_nif", "base_imponible",
        "tipo_iva", "cuota_iva", "total_factura", "porcentaje_retencion",
        "importe_retencion", "forma_pago", "issues", "confidence_score"
    ]
    
    try:
        # Verificar si hay múltiples tipos de IVA
        desglose_iva = data.get("desglose_iva", [])
        
        # Convertir issues de lista a string
        issues = data.get("issues", [])
        if isinstance(issues, list):
            data["issues"] = ", ".join(issues) if issues else ""
        
        if len(desglose_iva) <= 1:
            # Una sola línea
            flat_data = {col: data.get(col) for col in columns}
            
            # Extraer tipo de IVA si existe
            if desglose_iva and len(desglose_iva) > 0:
                flat_data["tipo_iva"] = desglose_iva[0].get("tipo_iva")
            
            # Preparar fila
            row = [_current_row_number] + [flat_data.get(col, "") for col in columns]
            _sheet.append(row)
            _current_row_number += 1
        else:
            # Múltiples tipos de IVA
            flat_data_principal = {col: data.get(col) for col in columns}
            
            primer_iva = desglose_iva[0]
            flat_data_principal["base_imponible"] = primer_iva.get("base_imponible")
            flat_data_principal["tipo_iva"] = primer_iva.get("tipo_iva")
            flat_data_principal["cuota_iva"] = primer_iva.get("cuota_iva")
            
            row_principal = [_current_row_number] + [flat_data_principal.get(col, "") for col in columns]
            _sheet.append(row_principal)
            _current_row_number += 1
            
            # Líneas adicionales
            for item_iva in desglose_iva[1:]:
                row_adicional = [
                    "", "", "", "", "", "",  # Vacías hasta base (6 vacías ahora)
                    item_iva.get("base_imponible"),
                    item_iva.get("tipo_iva"),
                    item_iva.get("cuota_iva"),
                    "", "", "", "", "", ""
                ]
                _sheet.append(row_adicional)
        
        print(f"   Guardado en Excel: {filename}")
        
    except Exception as e:
        print(f"   Error al guardar: {e}")

def finalizar_y_guardar_excel(filename: str = "facturas.xlsx"):
    """
    Ajusta columnas, aplica formato visual y guarda el archivo Excel.
    """
    global _workbook, _sheet
    
    if _workbook is None:
        return
    
    try:
        # Definir estilos
        light_blue_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        white_fill = PatternFill(fill_type=None)  # Sin relleno (blanco)
        border = Border(
            left=Side(style='thin', color='B4C7E7'),
            right=Side(style='thin', color='B4C7E7'),
            top=Side(style='thin', color='B4C7E7'),
            bottom=Side(style='thin', color='B4C7E7')
        )
        alignment_center = Alignment(horizontal="center", vertical="center")
        alignment_left = Alignment(horizontal="left", vertical="center")
        
        # Aplicar formato a filas de datos (desde fila 4 en adelante)
        max_row = _sheet.max_row
        max_col = _sheet.max_column
        
        # Agrupar filas por factura (detectando cuándo cambia el #)
        factura_actual = None
        color_actual = white_fill
        contador_facturas = 0
        
        for row_num in range(4, max_row + 1):
            primera_celda = _sheet.cell(row_num, 1).value
            es_total = primera_celda and str(primera_celda).upper() == "TOTAL"
            
            # Detectar si es una nueva factura (tiene número en columna #)
            if primera_celda and primera_celda != factura_actual and not es_total:
                factura_actual = primera_celda
                contador_facturas += 1
                # Alternar color por factura
                color_actual = light_blue_fill if contador_facturas % 2 == 0 else white_fill
            
            for col_num in range(1, max_col + 1):
                cell = _sheet.cell(row_num, col_num)
                
                # Aplicar bordes a todas las celdas
                cell.border = border
                
                if not es_total:
                    # Aplicar el color de la factura actual
                    cell.fill = color_actual
                    
                    # Alineación
                    if col_num == 1:  # Columna #
                        cell.alignment = alignment_center
                    elif col_num in [7, 8, 9, 10, 11, 12]:  # Columnas numéricas (ajustado por eliminar columna)
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = alignment_left
                else:
                    # Fila TOTAL: negrita y fondo amarillo claro
                    cell.font = Font(bold=True, size=11)
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                    cell.alignment = alignment_center if col_num == 1 else Alignment(horizontal="right", vertical="center")
        
        # Ajustar ancho de columnas
        for column in _sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value and not str(cell.value).startswith("="):
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            _sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Congelar paneles (primera fila y primeras 3 filas)
        _sheet.freeze_panes = "A4"
        
        _workbook.save(filename)
        print(f"\n   Excel guardado: {filename}")
        
        # Resetear variables globales
        _workbook = None
        _sheet = None
        
    except Exception as e:
        print(f"   Error al finalizar Excel: {e}")
