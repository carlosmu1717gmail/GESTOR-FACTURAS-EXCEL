# -*- coding: utf-8 -*-
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8') if hasattr(sys.stdout, 'reconfigure') else None

wb = openpyxl.load_workbook(r'facturas.xlsx')
ws = wb.active

print(f"Nombre hoja: {ws.title}")
print(f"Total filas: {ws.max_row}")
print(f"\nFila 1 (encabezados): {[ws.cell(1, i).value for i in range(1, 12)]}")
print(f"Fila 4 (primera factura): {[ws.cell(4, i).value for i in range(1, 12)]}")
print(f"Última fila: {[ws.cell(ws.max_row, i).value for i in range(1, 12)]}")

# Verificar si hay valores en las columnas de importes
print(f"\nBase Imponible (H4): {ws.cell(4, 8).value}")
print(f"% IVA (I4): {ws.cell(4, 9).value}")
print(f"Cuota IVA (J4): {ws.cell(4, 10).value}")
print(f"Total (K4): {ws.cell(4, 11).value}")
