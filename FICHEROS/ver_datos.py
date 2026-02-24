# -*- coding: utf-8 -*-
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8') if hasattr(sys.stdout, 'reconfigure') else None

wb = openpyxl.load_workbook(r'facturas01.xlsx')
ws = wb.active

print(f"=== Hoja: {ws.title} ===\n")

# Ver varias filas
for i in range(4, min(10, ws.max_row + 1)):
    print(f"Fila {i}:")
    print(f"  #: {ws.cell(i, 1).value}")
    print(f"  Archivo: {ws.cell(i, 2).value}")
    print(f"  Emisor: {ws.cell(i, 5).value}")
    print(f"  NIF: {ws.cell(i, 6).value}")
    print(f"  CP: {ws.cell(i, 7).value}")
    print(f"  Base: {ws.cell(i, 8).value}")
    print(f"  %IVA: {ws.cell(i, 9).value}")
    print(f"  Cuota: {ws.cell(i, 10).value}")
    print(f"  Total: {ws.cell(i, 11).value}")
    print()
