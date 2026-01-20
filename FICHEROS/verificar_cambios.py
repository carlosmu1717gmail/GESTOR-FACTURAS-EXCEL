# -*- coding: utf-8 -*-
import sys
import openpyxl
from openpyxl import load_workbook

# Configurar salida UTF-8
sys.stdout.reconfigure(encoding='utf-8') if hasattr(sys.stdout, 'reconfigure') else None

def verificar_excel(filename="facturas03.xlsx"):
    """Verificar que los cambios se aplicaron correctamente"""
    try:
        import os
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        file_path = os.path.join(base_dir, filename)
        wb = load_workbook(file_path)
        
        # Verificar hoja FACTURAS
        print("=" * 60)
        print("VERIFICACION DE CAMBIOS")
        print("=" * 60)
        
        sheet_facturas = wb["FACTURAS"]
        print("\n[FACTURAS] Hoja FACTURAS:")
        print("-" * 60)
        
        # Mostrar encabezados
        encabezados = []
        for col in range(1, sheet_facturas.max_column + 1):
            valor = sheet_facturas.cell(1, col).value
            if valor:
                encabezados.append(valor)
        
        print(f"Total de columnas: {len(encabezados)}")
        print("\nEncabezados:")
        for i, enc in enumerate(encabezados, 1):
            print(f"  {i}. {enc}")
        
        # Verificar que no existe "Tipo Documento"
        tiene_tipo_doc = any("Tipo Documento" in str(enc) for enc in encabezados)
        print(f"\n[OK] 'Tipo Documento' eliminado: {'NO [X]' if tiene_tipo_doc else 'SI [OK]'}")
        
        # Verificar que columnas G y H no tienen "Contraparte"
        col_g = sheet_facturas.cell(1, 5).value  # Columna 5 (E) ahora es Nombre
        col_h = sheet_facturas.cell(1, 6).value  # Columna 6 (F) ahora es NIF
        tiene_contraparte = "Contraparte" in str(col_g) or "Contraparte" in str(col_h)
        print(f"[OK] 'Contraparte' eliminado: {'NO [X]' if tiene_contraparte else 'SI [OK]'}")
        
        # Mostrar algunas fechas de ejemplo
        print("\nFormato de fechas (primeras 5 facturas):")
        for row in range(4, min(9, sheet_facturas.max_row + 1)):
            fecha = sheet_facturas.cell(row, 3).value  # Columna C (ahora Fecha Expedición)
            num_factura = sheet_facturas.cell(row, 4).value
            if fecha:
                tiene_puntos = '.' in str(fecha)
                print(f"  Fila {row}: {fecha} {'[!] (con puntos)' if tiene_puntos else '[OK]'}")
        
        # Verificar hoja PROVEEDORES
        print("\n" + "=" * 60)
        sheet_proveedores = wb["PROVEEDORES"]
        print("[PROVEEDORES] Hoja PROVEEDORES:")
        print("-" * 60)
        
        encabezados_prov = []
        for col in range(1, sheet_proveedores.max_column + 1):
            valor = sheet_proveedores.cell(1, col).value
            if valor:
                encabezados_prov.append(valor)
        
        print(f"Total de columnas: {len(encabezados_prov)}")
        print("\nEncabezados:")
        for i, enc in enumerate(encabezados_prov, 1):
            print(f"  {i}. {enc}")
        
        print("\n" + "=" * 60)
        print("VERIFICACION COMPLETADA")
        print("=" * 60)
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback

if __name__ == "__main__":
    verificar_excel()
