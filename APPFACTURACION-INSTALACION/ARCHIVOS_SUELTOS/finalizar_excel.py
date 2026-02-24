import openpyxl
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

def finalizar_excel(filename="facturas.xlsx"):
    """
    Añade totales generales y crea la hoja de PROVEEDORES.
    MEJORADO: Colores por proveedor, desglose de IVA por tipo, sin columna "Nombre Archivo".
    """
    try:
        workbook = load_workbook(filename)
        sheet = workbook.active
        
        # Renombrar la hoja "Sheet" a "FACTURAS"
        sheet.title = "FACTURAS"
        
        # 1. AÑADIR FILA DE TOTALES AL FINAL DE LA HOJA PRINCIPAL
        last_row = sheet.max_row
        total_row_number = last_row + 1
        
        # CORRECCIÓN: Usar SUM en lugar de SUMIF para incluir todas las líneas (incluso sin #)
        total_row = [
            "TOTAL",
            "", "", "", "", "",
            f"=SUM(G4:G{last_row})",  # Base Imponible
            "",
            f"=SUM(I4:I{last_row})",  # Cuota IVA
            f"=SUM(J4:J{last_row})",  # Total Factura
            "",
            f"=SUM(L4:L{last_row})",  # Retenciones
            "",
            "", ""
        ]
        
        sheet.append(total_row)
        
        for cell in sheet[total_row_number]:
            cell.font = Font(bold=True)
        
        # 2. CREAR HOJA DE PROVEEDORES (SIN COLUMNA "NOMBRE ARCHIVO")
        if "PROVEEDORES" in workbook.sheetnames:
            del workbook["PROVEEDORES"]
        
        sheet_proveedores = workbook.create_sheet("PROVEEDORES")
        
        # HEADERS SIN "Nombre Archivo"
        headers = ["#", "Fecha Expedición", "Número Factura", "Nombre", "NIF", 
                   "Base Imponible", "% IVA", "Cuota IVA", "Total Factura",
                   "%Retención", "Retenciones", "Forma de Pago",
                   "Observaciones", "Confianza"]
        sheet_proveedores.append(headers)
        
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in sheet_proveedores[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        sheet_proveedores.append([""] * len(headers))
        sheet_proveedores.append([""] * len(headers))
        
        # Leer datos principales
        datos = []
        for row_num in range(4, last_row + 1):
            numero_factura = sheet.cell(row=row_num, column=5).value
            if numero_factura:
                # EXTRAER SIN COLUMNA "NOMBRE ARCHIVO" (índice 1)
                row_data = []
                row_data.append(sheet.cell(row=row_num, column=1).value)  # #
                # SALTAR columna 2 (Nombre Archivo)
                for col_num in range(3, 16):  # Desde columna C hasta O
                    row_data.append(sheet.cell(row=row_num, column=col_num).value)
                datos.append((row_num, row_data))
        
        # Agrupar por NIF (ahora índice 4 sin "Nombre Archivo")
        proveedores = defaultdict(list)
        facturas_sin_nif = []
        
        for row_num, row_data in datos:
            nif = row_data[4]  # NIF ahora en índice 4
            if nif and str(nif).strip():
                proveedores[nif].append((row_num, row_data))
            else:
                facturas_sin_nif.append((row_num, row_data))
        
        # Paleta de colores para proveedores (colores suaves alternados)
        colores_proveedores = [
            "E8F4F8",  # Azul claro
            "FFF2CC",  # Amarillo claro
            "E2F0D9",  # Verde claro
            "FCE4D6",  # Naranja claro
            "F4E4F7",  # Morado claro
            "D9E1F2",  # Azul grisáceo
        ]
        
        current_row = 4
        inicio_datos = 4
        color_index = 0
        
        # ESCRIBIR PROVEEDORES CON COLOR Y DESGLOSE IVA
        for nif, facturas in sorted(proveedores.items()):
            color_proveedor = colores_proveedores[color_index % len(colores_proveedores)]
            fill_proveedor = PatternFill(start_color=color_proveedor, end_color=color_proveedor, fill_type="solid")
            color_index += 1
            
            start_row_proveedor = current_row
            
            # Diccionario para acumular por tipo de IVA
            desglose_iva_proveedor = defaultdict(lambda: {"base": 0, "cuota": 0, "total": 0})
            
            # Escribir facturas del proveedor
            for row_num_original, factura in facturas:
                sheet_proveedores.append(factura)
                
                # Aplicar color
                for col in range(1, len(headers) + 1):
                    sheet_proveedores.cell(row=current_row, column=col).fill = fill_proveedor
                
                # Acumular en desglose IVA
                tipo_iva = factura[6]  # % IVA (índice 6 sin Nombre Archivo)
                base_imp = factura[5] or 0  # Base Imponible
                cuota_iva = factura[7] or 0  # Cuota IVA
                total_fac = factura[8] or 0  # Total Factura
                
                if tipo_iva is not None and tipo_iva != "":
                    desglose_iva_proveedor[tipo_iva]["base"] += float(base_imp) if base_imp else 0
                    desglose_iva_proveedor[tipo_iva]["cuota"] += float(cuota_iva) if cuota_iva else 0
                    desglose_iva_proveedor[tipo_iva]["total"] += float(total_fac) if total_fac else 0
                
                current_row += 1
                
                # Copiar líneas adicionales de IVA
                siguiente_fila = row_num_original + 1
                while siguiente_fila <= last_row:
                    num_factura_siguiente = sheet.cell(row=siguiente_fila, column=5).value
                    if not num_factura_siguiente or not str(num_factura_siguiente).strip():
                        linea_adicional = []
                        linea_adicional.append(sheet.cell(row=siguiente_fila, column=1).value)
                        for col_num in range(3, 16):
                            linea_adicional.append(sheet.cell(row=siguiente_fila, column=col_num).value)
                        sheet_proveedores.append(linea_adicional)
                        
                        # Aplicar color
                        for col in range(1, len(headers) + 1):
                            sheet_proveedores.cell(row=current_row, column=col).fill = fill_proveedor
                        
                        # Acumular en desglose
                        tipo_iva_linea = linea_adicional[6]
                        base_linea = linea_adicional[5] or 0
                        cuota_linea = linea_adicional[7] or 0
                        total_linea = linea_adicional[8] or 0
                        
                        if tipo_iva_linea is not None and tipo_iva_linea != "":
                            desglose_iva_proveedor[tipo_iva_linea]["base"] += float(base_linea) if base_linea else 0
                            desglose_iva_proveedor[tipo_iva_linea]["cuota"] += float(cuota_linea) if cuota_linea else 0
                            desglose_iva_proveedor[tipo_iva_linea]["total"] += float(total_linea) if total_linea else 0
                        
                        current_row += 1
                        siguiente_fila += 1
                    else:
                        break
            
            end_row_proveedor = current_row - 1
            
            # AÑADIR SUBTOTALES POR TIPO DE IVA (OPCIÓN A)
            # SOLO SI HAY MÁS DE 1 TIPO DE IVA
            if len(desglose_iva_proveedor) > 1:
                for tipo_iva_key in sorted(desglose_iva_proveedor.keys(), reverse=True):
                    datos_iva = desglose_iva_proveedor[tipo_iva_key]
                    subtotal_iva_row = [
                        f"  IVA {tipo_iva_key}%",
                        "", "", "", "",
                        round(datos_iva["base"], 2),
                        tipo_iva_key,
                        round(datos_iva["cuota"], 2),
                        round(datos_iva["total"], 2),
                        "", "", "", "", ""
                    ]
                    sheet_proveedores.append(subtotal_iva_row)
                    
                    # Estilo para subtotales IVA (negrita, fondo más oscuro)
                    fill_subtotal_iva = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
                    for col in range(1, len(headers) + 1):
                        cell = sheet_proveedores.cell(row=current_row, column=col)
                        cell.fill = fill_subtotal_iva
                        cell.font = Font(bold=True, italic=True, size=10)
                    
                    current_row += 1
            
            # TOTAL GENERAL DEL PROVEEDOR (usando SUM para incluir TODO)
            subtotal_row = [
                f"TOTAL {nif}",
                "", "", "", "",
                f"=SUM(F{start_row_proveedor}:F{end_row_proveedor})",  # Base
                "",
                f"=SUM(H{start_row_proveedor}:H{end_row_proveedor})",  # Cuota IVA
                f"=SUM(I{start_row_proveedor}:I{end_row_proveedor})",  # Total
                "",
                f"=SUM(K{start_row_proveedor}:K{end_row_proveedor})",  # Retenciones
                "",
                "", ""
            ]
            
            sheet_proveedores.append(subtotal_row)
            
            # Estilo total proveedor (negrita, fondo dorado)
            fill_total = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
            for col in range(1, len(headers) + 1):
                cell = sheet_proveedores.cell(row=current_row, column=col)
                cell.fill = fill_total
                cell.font = Font(bold=True, size=11)
            
            current_row += 1
            
            # Línea en blanco
            sheet_proveedores.append([""] * len(headers))
            current_row += 1
        
        # GRUPO "SIN NIF" (mismo tratamiento)
        if facturas_sin_nif:
            fill_sin_nif = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            start_row_sin_nif = current_row
            
            for row_num_original, factura in facturas_sin_nif:
                sheet_proveedores.append(factura)
                for col in range(1, len(headers) + 1):
                    sheet_proveedores.cell(row=current_row, column=col).fill = fill_sin_nif
                current_row += 1
                
                # Copiar líneas adicionales
                siguiente_fila = row_num_original + 1
                while siguiente_fila <= last_row:
                    num_factura_siguiente = sheet.cell(row=siguiente_fila, column=5).value
                    if not num_factura_siguiente or not str(num_factura_siguiente).strip():
                        linea_adicional = []
                        linea_adicional.append(sheet.cell(row=siguiente_fila, column=1).value)
                        for col_num in range(3, 16):
                            linea_adicional.append(sheet.cell(row=siguiente_fila, column=col_num).value)
                        sheet_proveedores.append(linea_adicional)
                        for col in range(1, len(headers) + 1):
                            sheet_proveedores.cell(row=current_row, column=col).fill = fill_sin_nif
                        current_row += 1
                        siguiente_fila += 1
                    else:
                        break
            
            end_row_sin_nif = current_row - 1
            
            subtotal_sin_nif = [
                "TOTAL SIN NIF",
                "", "", "", "",
                f"=SUM(F{start_row_sin_nif}:F{end_row_sin_nif})",
                "",
                f"=SUM(H{start_row_sin_nif}:H{end_row_sin_nif})",
                f"=SUM(I{start_row_sin_nif}:I{end_row_sin_nif})",
                "",
                f"=SUM(K{start_row_sin_nif}:K{end_row_sin_nif})",
                "",
                "", ""
            ]
            
            sheet_proveedores.append(subtotal_sin_nif)
            
            for col in range(1, len(headers) + 1):
                cell = sheet_proveedores.cell(row=current_row, column=col)
                cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                cell.font = Font(bold=True)
            
            current_row += 1
            sheet_proveedores.append([""] * len(headers))
            current_row += 1
        
        fin_datos = current_row - 2
        
        # TOTAL GENERAL - CORRECCIÓN: Sumar SOLO los totales parciales (filas "TOTAL ...")
        # No sumar líneas individuales para evitar duplicación
        total_general_row = [
            "TOTAL GENERAL",
            "", "", "", "",
            f"=SUMIF(A{inicio_datos}:A{fin_datos},\"TOTAL*\",F{inicio_datos}:F{fin_datos})",
            "",
            f"=SUMIF(A{inicio_datos}:A{fin_datos},\"TOTAL*\",H{inicio_datos}:H{fin_datos})",
            f"=SUMIF(A{inicio_datos}:A{fin_datos},\"TOTAL*\",I{inicio_datos}:I{fin_datos})",
            "",
            f"=SUMIF(A{inicio_datos}:A{fin_datos},\"TOTAL*\",K{inicio_datos}:K{fin_datos})",
            "",
            "", ""
        ]
        
        sheet_proveedores.append(total_general_row)
        
        for col in range(1, len(headers) + 1):
            cell = sheet_proveedores.cell(row=current_row, column=col)
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=12)
        
        # APLICAR BORDES
        border = Border(
            left=Side(style='thin', color='B4C7E7'),
            right=Side(style='thin', color='B4C7E7'),
            top=Side(style='thin', color='B4C7E7'),
            bottom=Side(style='thin', color='B4C7E7')
        )
        
        for row in sheet_proveedores.iter_rows(min_row=1, max_row=current_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = border
        
        # Ajustar anchos de columna
        for sheet_obj in [sheet, sheet_proveedores]:
            for column in sheet_obj.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value and not str(cell.value).startswith("=SUM") and not str(cell.value).startswith("=FACTURAS"):
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                sheet_obj.column_dimensions[column_letter].width = adjusted_width
            
            sheet_obj.freeze_panes = "A4"
        
        workbook.save(filename)
        print(f"\n   Totales y hoja PROVEEDORES generados correctamente.")
        
    except Exception as e:
        print(f"\n   Error al finalizar Excel: {e}")
        import traceback
        traceback.print_exc()
