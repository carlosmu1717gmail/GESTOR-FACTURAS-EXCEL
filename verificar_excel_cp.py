import openpyxl
from openpyxl import load_workbook
import sys

# Configurar salida UTF-8
sys.stdout.reconfigure(encoding='utf-8') if hasattr(sys.stdout, 'reconfigure') else None

def verificar_cp(filename="facturas.xlsx"):
    """Verificar que la columna COD POSTAL existe y tiene datos"""
    try:
        if not os.path.exists(filename):
            print(f"ERROR: No se encuentra el archivo {filename}")
            return

        wb = load_workbook(filename)
        sheet = wb["FACTURAS"]
        
        print(f"\nVerificando archivo: {filename}")
        print("-" * 50)
        
        # Verificar encabezados
        headers = [cell.value for cell in sheet[1]]
        
        if "COD POSTAL" in headers:
            idx = headers.index("COD POSTAL") + 1
            print(f"[OK] Columna 'COD POSTAL' encontrada en posicion {idx}")
        else:
            print("[ERROR] Columna 'COD POSTAL' NO encontrada")
            return

        # Verificar datos (primeras 5 filas)
        print("\nMuestreo de datos (Primeras 5 facturas):")
        print(f"{'#':<5} | {'Nombre':<30} | {'CP':<10}")
        print("-" * 50)
        
        for row in range(4, min(sheet.max_row + 1, 9)):
            num = sheet.cell(row=row, column=1).value
            nombre = sheet.cell(row=row, column=5).value
            cp = sheet.cell(row=row, column=7).value # Columna 7 = G (COD POSTAL)
            
            if num:
                print(f"{str(num):<5} | {str(nombre)[:30]:<30} | {str(cp):<10}")
                
        print("-" * 50)
        print("Verificacion completada.")
        
    except Exception as e:
        print(f"ERROR durante verificacion: {e}")

if __name__ == "__main__":
    import os
    # Buscar el archivo más reciente que empiece por 'facturas'
    files = [f for f in os.listdir('.') if f.startswith('facturas') and f.endswith('.xlsx')]
    if files:
        newest = max(files, key=os.path.getctime)
        verificar_cp(newest)
    else:
        verificar_cp()
